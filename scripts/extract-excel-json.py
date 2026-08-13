"""Extract Stage Master + Rule Engine sheets to _excel_extract.json for generate-stage-master.py."""
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

excel = Path.home() / "Downloads" / "ZentroFlow_Consolidated_Developer_Specification (1).xlsx"
if not excel.exists():
    raise SystemExit(f"Excel not found: {excel}")

wb = openpyxl.load_workbook(excel, data_only=True)
out = {"stages": [], "rules": []}

ws = wb["02 Stage Master"]
headers = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
for row in ws.iter_rows(min_row=5, values_only=True):
    if not row[0]:
        continue
    out["stages"].append({headers[i]: ("" if row[i] is None else row[i]) for i in range(len(headers))})

ws = wb["03 Rule Engine"]
headers = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
for row in ws.iter_rows(min_row=5, values_only=True):
    if not row[0]:
        continue
    out["rules"].append({headers[i]: ("" if row[i] is None else row[i]) for i in range(len(headers))})

dest = Path(__file__).resolve().parents[1] / "src" / "domain" / "stages" / "_excel_extract.json"
dest.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
print("wrote", dest, "stages", len(out["stages"]), "rules", len(out["rules"]))
